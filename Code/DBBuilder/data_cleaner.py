#!/usr/bin/python
from openpyxl import Workbook
from openpyxl import load_workbook
from config import data_config

# data is an array: [handle, content, isRetweet, origAuthor, time, inReplyTo, isQuoteStatus, retweetCount, favCount, sourceURL]
def clean_row(data):
    # unpack columns needed
    handle, content, isRetweet, origAuthor, time = data[:5]
    retweetCount, favCount = [7:9]
    # convert encoded symbols, if any
    content = content.replace("&amp;", "&")
    content = content.replace("&lt;", "<")
    content = content.replace("&gt;", ">")
    # reformat time stamp
    time = time.replace("T"," ")
    # return only needed columns
    return [handle, content, time, isRetweet, origAuthor, retweetCount, favCount]

def clean_data():
    # open raw data file & get the worksheet
    input_file = data_config['input_filename']
    print '\nopening file:', input_file, '...'
    # a workbook is an xlsx file
    input_wb = load_workbook(input_file, read_only = True)
    # a worksheet is a sheet containing rows
    input_ws = input_wb[data_config['sheet_name']]

    # create new workbook with a new worksheet
    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = data_config['sheet_name']

    print 'cleaning data...'
    # for each row in input sheet
    for row in list(input_ws.rows):
        # extract cell data
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        # clean the data & add to new sheet
        output_ws.append(clean_row(row_data))

    # save wb
    output_wb.save(filename = data_config['clean_filename'])
    print 'DONE: saved file:', data_config['clean_filename']
